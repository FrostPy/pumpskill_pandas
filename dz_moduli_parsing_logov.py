import collections
from typing import Collection
from zlib import Z_HUFFMAN_ONLY
from numpy import number
import pandas
from  collections import Counter
from openpyxl import workbook
from openpyxl import load_workbook
from pprint import pprint

#report_output_file_name = 'report.xlsx'
#report_template_file_name = 'report_template.xlsx'
#log_file_name = 'logs.xlsx'
MOST_POPULAR_PRODUCT = 7
MOST_POPULAR_BROWSER = 7



def make_report(log_file_name , report_template_file_name = 'report_template.xlsx', report_output_file_name = 'report.xlsx'):
    # Чтение и анализ данных из Excel
    visit = []
    buys = []
    man_buy = []
    woman_buy = []
    vizit_browser = {}
    popular_month_buy = {}
    best_product = []
    
    log_file = pandas.read_excel(log_file_name, sheet_name='log', engine='openpyxl')
    log_file_dict = log_file.to_dict(orient='records')
    for slovar in log_file_dict:
        for key , value in slovar.items():
            if key == 'Браузер':
                visit.append(value)
            if key == 'Купленные товары':
                for i in value.split(','):
                    buys.append(i)
            if key == 'Пол' and value == 'м':
                for i in slovar['Купленные товары'].split(','):
                    man_buy.append(i) 
            if key == 'Пол' and value == 'ж': 
                for i in slovar['Купленные товары'].split(','):
                    woman_buy.append(i)  

        # Определяем самые популярные браузеры и покупаемые товары
    best_browser = collections.Counter(visit).most_common(MOST_POPULAR_BROWSER)
    best_product = collections.Counter(buys).most_common(MOST_POPULAR_PRODUCT)
    all_product_man = collections.Counter(man_buy).most_common()
    all_product_woman = collections.Counter(woman_buy).most_common()               
    most_popular_product_man = collections.Counter(man_buy).most_common(1)
    most_unpopular_product_man = all_product_man[:-(len(all_product_man) + 1):-1][0]
    most_popular_product_woman = collections.Counter(man_buy).most_common(1)
    most_unpopular_product_woman = all_product_woman[:-(len(all_product_woman) + 1):-1][0]

    for slovar in log_file_dict:
        datestamp = slovar['Дата посещения']
        date1 = datestamp.to_pydatetime()
        date2 = date1.date()
        number_month = int(date2.strftime("%m"))
        for i in range(MOST_POPULAR_BROWSER):
            if slovar['Браузер'] == str(best_browser[i][0]):
                if str(best_browser[i][0]) in vizit_browser:
                    for j in range(0,11):
                        if number_month == j:
                            if number_month in vizit_browser[str(best_browser[i][0])]:
                                vizit_browser[str(best_browser[i][0])][number_month] += 1
                            else:
                                vizit_browser[str(best_browser[i][0])][number_month]  = 1      
                else:
                    vizit_browser[str(best_browser[i][0])] = {number_month:1}
                

    for slovar in log_file_dict:
        datestamp = slovar['Дата посещения']
        date1 = datestamp.to_pydatetime()
        date2 = date1.date()
        number_month = int(date2.strftime("%m"))        
        
        temp_month = slovar['Купленные товары'].split(',')
        for elem in temp_month:
            for i in range(0, MOST_POPULAR_PRODUCT):
                for item in best_product[i]:
                    if elem == item:
                        if elem in popular_month_buy:
                            if number_month in popular_month_buy[elem]:
                                popular_month_buy[elem][number_month] += 1
                            else:
                                popular_month_buy[elem][number_month] = 1
                        else:
                            popular_month_buy[elem] = {}
                            popular_month_buy[elem][number_month] = 1                  

 # Открываем файл шаблона отчета report_template.xlsx
    wb = load_workbook(filename=report_template_file_name)
    ws = wb.active

    # Заполняем таблицу по использованию браузеров
    # В этом цикле заполняем популярные браузеры. Количество нормируется константой
    for i in range(1, MOST_POPULAR_BROWSER + 1):
        coordinats_column = 'A'
        res = coordinats_column + str(5 + i - 1)
        ws[res] = str(best_browser[i - 1][0])
        # Во вложенном цикле заполняем посещяемость по месяцам
        for j in range(1, 12):
            coordinats_column = ord('A')
            cell = chr(coordinats_column + j) + str(5 + i - 1)
            try:
                ws[cell] = str(vizit_browser[str(best_browser[i - 1][0])][j])
            except:
                pass

    # Заполняем таблицу по приобретенным товарам
    # В этом цикле заполняем популярные товаров. Количество нормируется константой
  
    for i in range(1, MOST_POPULAR_PRODUCT + 1):
        int_column = ord('A')
        row = chr(int_column) + str(19 + i - 1)
        ws[row] = str(best_product[i - 1][0])
        # pprint(most_popular_goods[i - 1][0])
        # Во вложенном цикле заполняем покупки популярных товаров по месяцам
        for j in range(1, 12):
            int_col = ord('A')
            cell = chr(int_col + j) + str(19 + i - 1)
            try:
                # Поиск ключа (товара, для которого нужно проставить количество продаж
                temp_key = str(best_product[i - 1][0])
                ws[cell] = str(popular_month_buy[temp_key][j])
            except:
                pass

        

    # Заполняем самые популярные и непопулярные товары у мужчин и женщин
    ws['B31'] = str(most_popular_product_man[0][0])
    ws['B32'] = str(most_popular_product_woman[0][0])
    ws['B33'] = str(most_unpopular_product_man[0])
    ws['B34'] = str(most_unpopular_product_woman[0])

    # Сохраняем файл-отчет
    wb.save(report_output_file_name)


make_report('logs.xlsx','report_template.xlsx','report.xlsx')